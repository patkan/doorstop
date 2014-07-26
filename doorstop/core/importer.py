"""Functions to import exiting documents and items."""

import os
import logging
import re
import csv

# TODO: track: openpyxl has false positives with pylint
# pylint: disable=E1101
import openpyxl  # pylint: disable=F0401
from openpyxl import load_workbook  # pylint: disable=F0401

from doorstop.common import DoorstopError, read_text, load_yaml
from doorstop.core.document import Document
from doorstop.core.item import Item
from doorstop.core.builder import _get_tree
from doorstop import settings

LIST_SEP_RE = re.compile(r"[\s;,]+")  # regex to split list strings into parts

_DOCUMENTS = []  # cache of unplaced documents


def import_file(path, document, ext=None, mapping=None, **kwargs):
    """Import items from an exported file.

    :param path: input file location
    :param document: document to import items
    :param ext: file extension to override input path's extension
    :param mapping: dictionary mapping custom to standard attribute names

    :raise DoorstopError: for unknown file formats

    :return: document with imported items

    """
    logging.info("importing {} into {}...".format(path, document))
    ext = ext or os.path.splitext(path)[-1]
    func = check(ext)
    func(path, document, mapping=mapping, **kwargs)


def create_document(prefix, path, parent=None, tree=None):
    """Create a Doorstop document from existing document information.

    :param prefix: existing document's prefix (for new items)
    :param path: new directory path to store this document's items
    :param parent: parent document's prefix (if one will exist)
    :param tree: explicit tree to add the document

    :return: imported Document

    """
    if not tree:
        tree = _get_tree()

    # Attempt to create a document with the given parent
    logging.info("importing document '{}'...".format(prefix))
    try:
        document = tree.create_document(path, prefix, parent=parent)
    except DoorstopError as exc:
        if not parent:
            raise exc from None

        # Create the document despite an unavailable parent
        document = Document.new(tree,
                                path, tree.root, prefix,
                                parent=parent)
        logging.warning(exc)
        _DOCUMENTS.append(document)

    # TODO: attempt to place unplaced documents?

    # Cache and return the document
    logging.info("imported: {}".format(document))
    if settings.CACHE_DOCUMENTS:
        tree._document_cache[document.prefix] = document  # pylint: disable=W0212
    return document


def add_item(prefix, identifier, attrs=None, document=None):
    """Create a Doorstop document from existing document information.

    :param prefix: previously imported document's prefix
    :param identifier: existing item's unique ID
    :param attrs: dictionary of Doorstop and custom attributes
    :param document: explicit document to add the item

    :return: imported Item

    """
    if document:
        # Get an explicit tree
        tree = document.tree
        assert tree  # tree should be set internally
    else:
        # Get an implicit tree and document
        tree = _get_tree()
        document = tree.find_document(prefix)

    # Add an item using the specified identifier
    logging.info("importing item '{}'...".format(identifier))
    item = Item.new(tree, document,
                    document.path, document.root, identifier,
                    auto=False)
    for key, value in (attrs or {}).items():
        item.set(key, value)
    item.save()

    logging.info("imported: {}".format(item))
    return item


def _file_yml(path, document, **_):
    """Import items from a YAML export to a document.

    :param path: input file location
    :param document: document to import items

    """
    # Parse the file
    logging.info("reading items in {}...".format(path))
    text = read_text(path)
    # Load the YAML data
    data = load_yaml(text, path)
    # Add items
    for identifier, attrs in data.items():
        try:
            item = document.find_item(identifier)
        except DoorstopError:
            pass
        else:
            item.delete()
        add_item(document.prefix, identifier, attrs=attrs, document=document)


def _file_csv(path, document, delimiter=',', mapping=None):
    """Import items from a CSV export to a document.

    :param path: input file location
    :param document: document to import items
    :param delimiter: CSV field delimiter
    :param mapping: dictionary mapping custom to standard attribute names

    """
    rows = []

    # Parse the file
    logging.info("reading rows in {}...".format(path))
    with open(path, 'r', encoding='utf-8') as stream:
        reader = csv.reader(stream, delimiter=delimiter)
        for _row in reader:
            row = []
            for value in _row:
                # convert string booleans
                if isinstance(value, str):
                    if value.lower() == 'true':
                        value = True
                    elif value.lower() == 'false':
                        value = False
                row.append(value)
            rows.append(row)

    # Extract header and data rows
    header = rows[0]
    data = rows[1:]

    # Import items from the rows
    _itemize(header, data, document, mapping=mapping)


def _file_tsv(path, document, mapping=None):
    """Import items from a TSV export to a document.

    :param path: input file location
    :param document: document to import items
    :param mapping: dictionary mapping custom to standard attribute names

    """
    _file_csv(path, document, delimiter='\t', mapping=mapping)


def _file_xlsx(path, document, mapping=None):
    """Import items from an XLSX export to a document.

    :param path: input file location
    :param document: document to import items
    :param mapping: dictionary mapping custom to standard attribute names

    """
    header = []
    data = []

    # Parse the file
    logging.debug("reading rows in {}...".format(path))
    workbook = load_workbook(path)
    worksheet = workbook.active

    # Locate the bottom right cell in the workbook that contains cell info
    _highest_column = worksheet.get_highest_column()
    _highest_letter = openpyxl.cell.get_column_letter(_highest_column)
    _highest_row = worksheet.get_highest_row()
    last_cell = _highest_letter + str(_highest_row)

    # Extract header and data rows
    for index, row in enumerate(worksheet.range('A1:%s' % last_cell)):
        row2 = []
        for cell in row:
            if index == 0:
                header.append(cell.value)
            else:
                row2.append(cell.value)
        if index:
            data.append(row2)

    # Import items from the rows
    _itemize(header, data, document, mapping=mapping)


def _itemize(header, data, document, mapping=None):
    """Conversion function for multiple formats.

    :param header: list of columns names
    :param data: list of lists of row values
    :param document: document to import items
    :param mapping: dictionary mapping custom to standard attribute names

    """
    logging.info("converting rows to items...")
    logging.debug("header: {}".format(header))
    for row in data:
        logging.debug("row: {}".format(row))

        # Parse item attributes
        attrs = {}
        identifier = None
        for index, value in enumerate(row):

            # Key lookup
            key = str(header[index]).lower().strip() if header[index] else ''
            if not key:
                continue

            # Map key to custom attributes names
            for custom, standard in (mapping or {}).items():
                if key == custom.lower():
                    msg = "mapped: '{}' => '{}'".format(key, standard)
                    logging.debug(msg)
                    key = standard
                    break

            # Convert values for particular keys
            if key == 'id':
                identifier = value
            elif key == 'links':
                # split links into a list
                attrs[key] = _split_list(value)
            else:
                attrs[key] = value

        # Convert the row to an item
        if identifier:

            # Delete the old item
            try:
                item = document.find_item(identifier)
            except DoorstopError:
                logging.debug("not yet an item: {}".format(identifier))
            else:
                logging.debug("deleting old item: {}".format(identifier))
                item.delete()

            # Import the item
            try:
                item = add_item(document.prefix, identifier,
                                attrs=attrs, document=document)
            except DoorstopError as exc:
                logging.warning(exc)


def _split_list(value):
    """Split a string list into parts."""
    if value:
        return [p for p in LIST_SEP_RE.split(value) if p]
    else:
        return []


# Mapping from file extension to file reader
FORMAT_FILE = {'.yml': _file_yml,
               '.csv': _file_csv,
               '.tsv': _file_tsv,
               '.xlsx': _file_xlsx}


def check(ext):
    """Confirm an extension is supported for import.

    :raise DoorstopError: for unknown formats

    :return: file importer if available

    """
    exts = ', '.join(ext for ext in FORMAT_FILE)
    msg = "unknown import format: {} (options: {})".format(ext or None, exts)
    exc = DoorstopError(msg)
    try:
        func = FORMAT_FILE[ext]
    except KeyError:
        raise exc from None
    else:
        logging.debug("found file reader for: {}".format(ext))
        return func
